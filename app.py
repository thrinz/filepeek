import getpass
import hashlib
import hmac
import json
import mimetypes
import os
import secrets
import shutil
import sys
import tempfile
import threading
import time
import uuid
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from html import escape as html_escape
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, Query, Request, UploadFile, File, Form
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from starlette.background import BackgroundTask

ROOT = Path(os.environ.get("FILEPEEK_ROOT", str(Path.home()))).expanduser().resolve()
STATIC_DIR = Path(__file__).parent / "static"
STATE_DIR = Path(os.environ.get("FILEPEEK_STATE_DIR", str(Path(__file__).parent))).expanduser().resolve()
PERMLINKS_FILE = STATE_DIR / "permlinks.json"
BOOKMARKS_FILE = STATE_DIR / "bookmarks.json"
HTML_EXTS = {".html", ".htm"}
MD_EXTS = {".md", ".markdown"}
PERMLINK_EXTS = HTML_EXTS | MD_EXTS  # file types that can be perma-linked / rendered at /view

TEXT_EXTS = {
    ".txt", ".md", ".markdown", ".rst", ".log", ".json", ".yaml", ".yml",
    ".toml", ".ini", ".cfg", ".conf", ".env", ".sh", ".bash", ".zsh",
    ".py", ".pyx", ".pyi", ".js", ".jsx", ".ts", ".tsx", ".mjs", ".cjs",
    ".html", ".htm", ".css", ".scss", ".sass", ".less", ".xml", ".svg",
    ".csv", ".tsv", ".sql", ".go", ".rs", ".java", ".kt", ".c", ".h",
    ".cpp", ".hpp", ".cs", ".rb", ".php", ".pl", ".lua", ".r", ".swift",
    ".vue", ".dart", ".gradle", ".properties", ".gitignore", ".dockerfile",
    ".tf", ".hcl", ".proto", ".mmd", ".mermaid",
}
SHEET_EXTS = {".xlsx", ".xlsm"}
DOC_EXTS = {".docx"}
PPT_EXTS = {".pptx"}
MAX_SHEET_ROWS = 2000      # rows rendered per worksheet
MAX_SHEET_COLS = 100       # columns rendered per worksheet
MAX_OFFICE_BYTES = 25 * 1024 * 1024  # don't try to render office files larger than this

BINARY_BYTE_LIMIT = 8192
MAX_TEXT_BYTES = 16 * 1024 * 1024  # 16MB inline editor cap
SEARCH_FILE_LIMIT = 1 * 1024 * 1024  # 1MB per file for content search
MAX_TRANSFER_BYTES = 500 * 1024 * 1024  # 500MB upload/download cap
UPLOAD_CHUNK = 1024 * 1024  # 1MB streaming chunks

# --- Authentication (remote mode) ---------------------------------------
# Auth turns on when FILEPEEK_PASSWORD_HASH and/or FILEPEEK_TOKEN is set.
# Generate a hash with:  python app.py hash-password
PASSWORD_HASH = os.environ.get("FILEPEEK_PASSWORD_HASH", "")
API_TOKEN = os.environ.get("FILEPEEK_TOKEN", "")
AUTH_ENABLED = bool(PASSWORD_HASH or API_TOKEN)
# Unset secret means sessions are invalidated on restart, which is fine for local use
SESSION_SECRET = os.environ.get("FILEPEEK_SECRET") or secrets.token_hex(32)
SESSION_COOKIE = "filepeek_session"
SESSION_TTL = 7 * 24 * 3600
LOGIN_MAX_FAILURES = 10
LOGIN_LOCKOUT_SECONDS = 15 * 60
AUTH_EXEMPT_PATHS = {"/login", "/static/logo.svg"}

_login_failures: dict = {}  # ip -> (failure_count, locked_until_ts)


def hash_password(password: str, iterations: int = 200_000) -> str:
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), iterations)
    return f"pbkdf2_sha256${iterations}${salt}${digest.hex()}"


def verify_password(password: str, stored: str) -> bool:
    try:
        algo, iterations, salt, expected = stored.split("$")
        if algo != "pbkdf2_sha256":
            return False
        digest = hashlib.pbkdf2_hmac("sha256", password.encode(), bytes.fromhex(salt), int(iterations))
        return hmac.compare_digest(digest.hex(), expected)
    except (ValueError, TypeError):
        return False


def _sign_session(expiry: int) -> str:
    sig = hmac.new(SESSION_SECRET.encode(), str(expiry).encode(), hashlib.sha256).hexdigest()
    return f"{expiry}.{sig}"


def _session_valid(cookie: Optional[str]) -> bool:
    if not cookie or "." not in cookie:
        return False
    expiry, _, sig = cookie.partition(".")
    try:
        if int(expiry) < time.time():
            return False
    except ValueError:
        return False
    expected = hmac.new(SESSION_SECRET.encode(), expiry.encode(), hashlib.sha256).hexdigest()
    return hmac.compare_digest(sig, expected)


def _client_ip(request: Request) -> str:
    # Trust X-Forwarded-For only when the direct peer is the local reverse proxy
    direct = request.client.host if request.client else "unknown"
    fwd = request.headers.get("x-forwarded-for")
    if fwd and direct in ("127.0.0.1", "::1"):
        return fwd.split(",")[0].strip()
    return direct


app = FastAPI(title="filepeek")


@app.middleware("http")
async def require_auth(request, call_next):
    if AUTH_ENABLED and request.url.path not in AUTH_EXEMPT_PATHS:
        authorized = _session_valid(request.cookies.get(SESSION_COOKIE))
        if not authorized and API_TOKEN:
            header = request.headers.get("authorization", "")
            authorized = header.startswith("Bearer ") and hmac.compare_digest(header[7:], API_TOKEN)
        if not authorized:
            if "text/html" in request.headers.get("accept", ""):
                return RedirectResponse("/login", status_code=303)
            return JSONResponse({"detail": "Not authenticated"}, status_code=401)
    return await call_next(request)


@app.middleware("http")
async def cap_request_size(request, call_next):
    cl = request.headers.get("content-length")
    if cl is not None:
        try:
            if int(cl) > MAX_TRANSFER_BYTES:
                return JSONResponse(
                    {"detail": f"Request body exceeds {MAX_TRANSFER_BYTES // (1024*1024)}MB limit"},
                    status_code=413,
                )
        except ValueError:
            pass
    return await call_next(request)


def safe_path(rel: str) -> Path:
    rel = (rel or "").lstrip("/")
    resolved = (ROOT / rel).resolve()
    if resolved != ROOT and ROOT not in resolved.parents:
        raise HTTPException(status_code=403, detail="Path outside root")
    return resolved


def to_rel(p: Path) -> str:
    try:
        rel = str(p.relative_to(ROOT))
        return "" if rel == "." else rel
    except ValueError:
        return ""


def is_text_file(p: Path) -> bool:
    if p.suffix.lower() in TEXT_EXTS:
        return True
    try:
        with p.open("rb") as f:
            chunk = f.read(BINARY_BYTE_LIMIT)
    except OSError:
        return False
    if not chunk:
        return True
    if b"\x00" in chunk:
        return False
    try:
        chunk.decode("utf-8")
        return True
    except UnicodeDecodeError:
        return False


def read_sheet(p: Path) -> list:
    """Read an .xlsx/.xlsm workbook into a list of {name, rows, truncated} sheets."""
    from openpyxl import load_workbook

    wb = load_workbook(p, read_only=True, data_only=True)
    try:
        sheets = []
        for ws in wb.worksheets:
            rows, truncated = [], False
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= MAX_SHEET_ROWS:
                    truncated = True
                    break
                rows.append(["" if v is None else str(v) for v in row[:MAX_SHEET_COLS]])
            # pad ragged rows to a uniform column count
            ncols = max((len(r) for r in rows), default=0)
            for r in rows:
                if len(r) < ncols:
                    r.extend([""] * (ncols - len(r)))
            sheets.append({"name": ws.title, "rows": rows, "truncated": truncated})
        return sheets
    finally:
        wb.close()


W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"


def _toggle_on(rPr, tag: str) -> bool:
    """True if a run-property toggle (e.g. b, i) is present and not explicitly off."""
    el = rPr.find(f"{W}{tag}")
    return el is not None and (el.get(f"{W}val") not in ("false", "0", "none"))


def _runs_html(p) -> str:
    """Render a paragraph's runs, preserving bold/italic as <strong>/<em>."""
    out = []
    for r in p.findall(f"{W}r"):
        text = "".join(t.text or "" for t in r.iter(f"{W}t"))
        if not text:
            continue
        frag = html_escape(text)
        rPr = r.find(f"{W}rPr")
        if rPr is not None:
            if _toggle_on(rPr, "b"):
                frag = f"<strong>{frag}</strong>"
            if _toggle_on(rPr, "i"):
                frag = f"<em>{frag}</em>"
        out.append(frag)
    return "".join(out)


def _docx_block_html(el) -> str:
    """Render a docx body block (paragraph or table) as minimal HTML."""
    if el.tag == f"{W}p":
        text = "".join(t.text or "" for t in el.iter(f"{W}t"))
        style = ""
        pPr = el.find(f"{W}pPr")
        if pPr is not None:
            pStyle = pPr.find(f"{W}pStyle")
            if pStyle is not None:
                style = (pStyle.get(f"{W}val") or "").lower()
        if not text.strip():
            return "<p><br></p>"
        if style == "title":
            return f"<h1>{html_escape(text)}</h1>"
        if style.startswith("heading"):
            digits = "".join(c for c in style if c.isdigit())
            lvl = min(int(digits), 6) if digits else 2
            return f"<h{lvl}>{html_escape(text)}</h{lvl}>"
        return f"<p>{_runs_html(el)}</p>"
    if el.tag == f"{W}tbl":
        out = ['<table>']
        for tr in el.findall(f"{W}tr"):
            out.append("<tr>")
            for tc in tr.findall(f"{W}tc"):
                inner = "".join(
                    _docx_block_html(c) for c in tc if c.tag in (f"{W}p", f"{W}tbl")
                )
                out.append(f"<td>{inner}</td>")
            out.append("</tr>")
        out.append("</table>")
        return "".join(out)
    return ""


def read_doc_html(p: Path) -> str:
    """Extract a .docx document's text/tables as minimal HTML (stdlib only)."""
    with zipfile.ZipFile(p) as z:
        xml = z.read("word/document.xml")
    body = ET.fromstring(xml).find(f"{W}body")
    if body is None:
        return ""
    return "".join(_docx_block_html(el) for el in body)


A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


def read_pptx_html(p: Path) -> str:
    """Extract a .pptx presentation's slide text as minimal HTML, one block per slide.

    Slides are ordered by the trailing number in their filename (slide1.xml,
    slide2.xml, …). Within a slide, each non-empty drawingml paragraph (a:p)
    becomes a line; the first text block is treated as the slide title.
    """
    def slide_num(name: str) -> int:
        digits = "".join(c for c in name[len("ppt/slides/slide"):] if c.isdigit())
        return int(digits) if digits else 0

    with zipfile.ZipFile(p) as z:
        names = sorted(
            (n for n in z.namelist()
             if n.startswith("ppt/slides/slide") and n.endswith(".xml")),
            key=slide_num,
        )
        out = []
        for i, name in enumerate(names, 1):
            root = ET.fromstring(z.read(name))
            lines = []
            for para in root.iter(f"{A}p"):
                text = "".join(t.text or "" for t in para.iter(f"{A}t"))
                if text.strip():
                    lines.append(text)
            body = []
            if lines:
                body.append(f"<h2>{html_escape(lines[0])}</h2>")
                body.extend(f"<p>{html_escape(line)}</p>" for line in lines[1:])
            else:
                body.append('<p class="pptx-empty">(no text on this slide)</p>')
            out.append(
                f'<section class="pptx-slide">'
                f'<div class="pptx-slide-num">Slide {i}</div>{"".join(body)}</section>'
            )
    if not out:
        return '<div class="pptx-empty">No slides found.</div>'
    return "".join(out)


def file_info(p: Path) -> dict:
    st = p.stat()
    return {
        "name": p.name,
        "path": to_rel(p),
        "is_dir": p.is_dir(),
        "size": st.st_size if not p.is_dir() else 0,
        "modified": datetime.fromtimestamp(st.st_mtime).isoformat(),
        "modified_ts": st.st_mtime,
    }


@app.get("/", response_class=HTMLResponse)
def index():
    return (STATIC_DIR / "index.html").read_text().replace("__FILEPEEK_ROOT__", str(ROOT))


LOGIN_PAGE = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>filepeek — log in</title>
<link rel="icon" type="image/svg+xml" href="/static/logo.svg">
<style>
  /* family "ink" tokens — keep in sync with webterm's static/style.css :root palette */
  :root {
    --bg: #1e293b; --card: #fff; --fg: #1e293b; --border: #cbd5e1;
    --accent: #2563eb; --accent-hover: #1d4ed8; --danger: #dc2626;
  }
  @media (prefers-color-scheme: dark) {
    :root {
      --bg: #0b0e14; --card: #11151f; --fg: #d7dce5; --border: #232a3a;
      --accent: #2563eb; --accent-hover: #1d4ed8; --danger: #e05c5c;
    }
  }
  body { margin: 0; min-height: 100vh; display: flex; align-items: center; justify-content: center;
         background: var(--bg); font-family: system-ui, sans-serif; }
  .card { background: var(--card); border-radius: 12px; padding: 2rem 2.5rem; width: 20rem;
          box-shadow: 0 10px 30px rgba(0,0,0,.3); text-align: center; }
  .card img { width: 48px; height: 48px; }
  h1 { font-size: 1.25rem; margin: .75rem 0 1.25rem; color: var(--fg); }
  input { width: 100%; box-sizing: border-box; padding: .6rem .75rem; border: 1px solid var(--border);
          border-radius: 8px; font-size: 1rem; background: var(--card); color: var(--fg); }
  button { width: 100%; margin-top: .75rem; padding: .6rem; border: 0; border-radius: 8px;
           background: var(--accent); color: #fff; font-size: 1rem; cursor: pointer; }
  button:hover { background: var(--accent-hover); }
  .err { color: var(--danger); font-size: .875rem; min-height: 1.25rem; margin: .5rem 0 0; }
</style>
</head>
<body>
  <form class="card" method="post" action="/login">
    <img src="/static/logo.svg" alt="">
    <h1>filepeek</h1>
    <input type="password" name="password" placeholder="Password" autofocus autocomplete="current-password">
    <p class="err">__ERROR__</p>
    <button type="submit">Log in</button>
  </form>
</body>
</html>"""


@app.get("/login", response_class=HTMLResponse)
def login_page():
    if not AUTH_ENABLED:
        return RedirectResponse("/", status_code=303)
    return LOGIN_PAGE.replace("__ERROR__", "")


@app.post("/login")
def login(request: Request, password: str = Form("")):
    if not AUTH_ENABLED:
        return RedirectResponse("/", status_code=303)
    ip = _client_ip(request)
    now = time.time()
    count, locked_until = _login_failures.get(ip, (0, 0.0))
    if now < locked_until:
        return HTMLResponse(
            LOGIN_PAGE.replace("__ERROR__", "Too many attempts — try again in a few minutes."),
            status_code=429,
        )
    if not (PASSWORD_HASH and verify_password(password, PASSWORD_HASH)):
        time.sleep(0.5)  # slow down brute force; sync handler so the event loop is unaffected
        count += 1
        locked = now + LOGIN_LOCKOUT_SECONDS if count >= LOGIN_MAX_FAILURES else 0.0
        _login_failures[ip] = (count, locked)
        return HTMLResponse(LOGIN_PAGE.replace("__ERROR__", "Wrong password."), status_code=401)
    _login_failures.pop(ip, None)
    resp = RedirectResponse("/", status_code=303)
    secure = request.url.scheme == "https" or request.headers.get("x-forwarded-proto") == "https"
    resp.set_cookie(
        SESSION_COOKIE, _sign_session(int(now) + SESSION_TTL),
        max_age=SESSION_TTL, httponly=True, samesite="lax", secure=secure,
    )
    return resp


@app.post("/logout")
def logout():
    resp = RedirectResponse("/login", status_code=303)
    resp.delete_cookie(SESSION_COOKIE)
    return resp


@app.get("/api/tree")
def list_dir(path: str = "", sort: str = "name"):
    target = safe_path(path)
    if not target.exists():
        raise HTTPException(404, "Not found")
    if not target.is_dir():
        raise HTTPException(400, "Not a directory")

    entries = []
    for child in target.iterdir():
        try:
            entries.append(file_info(child))
        except OSError:
            continue

    if sort == "modified":
        entries.sort(key=lambda e: (not e["is_dir"], -e["modified_ts"]))
    else:
        entries.sort(key=lambda e: (not e["is_dir"], e["name"].lower()))

    return {"path": to_rel(target), "entries": entries}


@app.get("/api/file")
def get_file(path: str):
    p = safe_path(path)
    if not p.exists():
        raise HTTPException(404, "Not found")
    if p.is_dir():
        raise HTTPException(400, "Is a directory")

    st = p.stat()
    mime, _ = mimetypes.guess_type(p.name)
    info = {
        "path": to_rel(p),
        "name": p.name,
        "size": st.st_size,
        "modified": datetime.fromtimestamp(st.st_mtime).isoformat(),
        "mime": mime or "application/octet-stream",
        "ext": p.suffix.lower(),
    }
    ext = p.suffix.lower()
    if ext in SHEET_EXTS or ext in DOC_EXTS or ext in PPT_EXTS:
        info["is_text"] = False
        info["content"] = None
        if st.st_size > MAX_OFFICE_BYTES:
            info["office_error"] = f"File exceeds {MAX_OFFICE_BYTES // (1024*1024)}MB render limit"
        else:
            try:
                if ext in SHEET_EXTS:
                    info["office"] = {"type": "sheet", "sheets": read_sheet(p)}
                elif ext in PPT_EXTS:
                    info["office"] = {"type": "pptx", "html": read_pptx_html(p)}
                else:
                    info["office"] = {"type": "doc", "html": read_doc_html(p)}
            except Exception as e:  # noqa: BLE001 - surface any parse failure to the client
                info["office_error"] = str(e)
    elif is_text_file(p) and st.st_size <= MAX_TEXT_BYTES:
        try:
            info["is_text"] = True
            info["content"] = p.read_text(encoding="utf-8", errors="replace")
        except OSError as e:
            raise HTTPException(500, str(e))
    else:
        info["is_text"] = False
        info["content"] = None
    return info


class CreateBody(BaseModel):
    path: str
    is_dir: bool = False


@app.post("/api/create")
def create_item(body: CreateBody):
    p = safe_path(body.path)
    if p.exists():
        raise HTTPException(409, "Already exists")
    p.parent.mkdir(parents=True, exist_ok=True)
    if body.is_dir:
        p.mkdir()
    else:
        p.touch()
    return file_info(p)


class RenameBody(BaseModel):
    path: str
    name: str


@app.post("/api/rename")
def rename_item(body: RenameBody):
    src = safe_path(body.path)
    if not src.exists():
        raise HTTPException(404, "Not found")
    new_name = body.name.strip()
    if not new_name or "/" in new_name or "\\" in new_name or new_name in (".", ".."):
        raise HTTPException(400, "Invalid name")
    dest = safe_path(str((src.parent / new_name).relative_to(ROOT)))
    if dest == src:
        return file_info(src)
    if dest.exists():
        raise HTTPException(409, "A file with that name already exists")
    src.rename(dest)
    return file_info(dest)


@app.delete("/api/delete")
def delete_item(path: str):
    p = safe_path(path)
    if p == ROOT:
        raise HTTPException(403, "Cannot delete the root folder")
    if not p.exists():
        raise HTTPException(404, "Not found")
    if p.is_dir():
        shutil.rmtree(p)
    else:
        p.unlink()
    # Drop any permlinks/bookmarks that pointed at the deleted path (or its descendants).
    rel = to_rel(p)
    prefix = rel + "/"
    links = _load_permlinks()
    kept = [l for l in links if l["path"] != rel and not l["path"].startswith(prefix)]
    if len(kept) != len(links):
        _save_permlinks(kept)
    marks = _load_bookmarks()
    kept_marks = [m for m in marks if m["path"] != rel and not m["path"].startswith(prefix)]
    if len(kept_marks) != len(marks):
        _save_bookmarks(kept_marks)
    return {"ok": True}


class TransferBody(BaseModel):
    path: str          # source file or folder (relative)
    dest_dir: str      # destination directory (relative; "" = root)
    overwrite: bool = False


def _prepare_transfer(body: TransferBody) -> tuple:
    src = safe_path(body.path)
    if src == ROOT:
        raise HTTPException(403, "Cannot transfer the root folder")
    if not src.exists():
        raise HTTPException(404, "Source not found")
    dest_dir = safe_path(body.dest_dir)
    if not dest_dir.exists() or not dest_dir.is_dir():
        raise HTTPException(404, "Destination directory not found")
    dest = (dest_dir / src.name).resolve()
    if dest != ROOT and ROOT not in dest.parents:
        raise HTTPException(403, "Destination outside root")
    if dest == src:
        raise HTTPException(400, "Source and destination are the same")
    # Block copying/moving a folder into itself or one of its own descendants.
    if src.is_dir() and (dest_dir == src or src in dest_dir.parents):
        raise HTTPException(400, "Cannot place a folder inside itself")
    if dest.exists() and not body.overwrite:
        raise HTTPException(409, "An item with that name already exists at the destination")
    return src, dest


def _remove_existing(dest: Path) -> None:
    if dest.is_dir():
        shutil.rmtree(dest)
    else:
        dest.unlink()


@app.post("/api/copy")
def copy_item(body: TransferBody):
    src, dest = _prepare_transfer(body)
    if dest.exists():
        _remove_existing(dest)
    if src.is_dir():
        shutil.copytree(src, dest)
    else:
        shutil.copy2(src, dest)
    return file_info(dest)


@app.post("/api/move")
def move_item(body: TransferBody):
    src, dest = _prepare_transfer(body)
    if dest.exists():
        _remove_existing(dest)
    shutil.move(str(src), str(dest))
    return file_info(dest)


class SaveBody(BaseModel):
    path: str
    content: str


@app.put("/api/file")
def save_file(body: SaveBody):
    p = safe_path(body.path)
    if p.is_dir():
        raise HTTPException(400, "Is a directory")
    if len(body.content.encode("utf-8")) > MAX_TRANSFER_BYTES:
        raise HTTPException(413, f"Content exceeds {MAX_TRANSFER_BYTES} bytes")
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(body.content, encoding="utf-8")
    return file_info(p)


@app.post("/api/upload")
async def upload_file(
    file: UploadFile = File(...),
    target_dir: str = Form(""),
    overwrite: bool = Form(False),
):
    dest_dir = safe_path(target_dir)
    if not dest_dir.exists():
        raise HTTPException(404, "Target directory does not exist")
    if not dest_dir.is_dir():
        raise HTTPException(400, "Target is not a directory")

    dest = safe_path(f"{target_dir}/{file.filename}" if target_dir else file.filename)
    if dest.exists() and not overwrite:
        raise HTTPException(409, "File already exists")

    written = 0
    try:
        with dest.open("wb") as out:
            while True:
                chunk = await file.read(UPLOAD_CHUNK)
                if not chunk:
                    break
                written += len(chunk)
                if written > MAX_TRANSFER_BYTES:
                    out.close()
                    dest.unlink(missing_ok=True)
                    raise HTTPException(413, f"Upload exceeds {MAX_TRANSFER_BYTES // (1024*1024)}MB limit")
                out.write(chunk)
    finally:
        await file.close()
    return file_info(dest)


@app.get("/api/download")
def download(path: str):
    p = safe_path(path)
    if not p.exists() or p.is_dir():
        raise HTTPException(404, "Not found")
    if p.stat().st_size > MAX_TRANSFER_BYTES:
        raise HTTPException(413, f"File exceeds {MAX_TRANSFER_BYTES // (1024*1024)}MB download limit")
    return FileResponse(p, filename=p.name)


# --- Folder zip download (job-based so the client can poll progress) -----
ZIP_JOBS: dict = {}
ZIP_JOB_TTL = 3600  # seconds before an unclaimed finished job is reaped


def _reap_zip_jobs() -> None:
    now = time.time()
    for jid, job in list(ZIP_JOBS.items()):
        if job["status"] != "running" and now - job["created"] > ZIP_JOB_TTL:
            try:
                os.unlink(job["file"])
            except OSError:
                pass
            ZIP_JOBS.pop(jid, None)


class ZipBody(BaseModel):
    path: str


@app.post("/api/zip")
def start_zip(body: ZipBody):
    target = safe_path(body.path)
    if not target.exists() or not target.is_dir():
        raise HTTPException(404, "Folder not found")

    files, total_bytes = [], 0
    for dirpath, dirnames, filenames in os.walk(target):
        for name in filenames:
            fp = Path(dirpath) / name
            if fp.is_symlink():
                continue
            try:
                size = fp.stat().st_size
            except OSError:
                continue
            files.append((fp, size))
            total_bytes += size
    if total_bytes > MAX_TRANSFER_BYTES:
        raise HTTPException(413, f"Folder exceeds {MAX_TRANSFER_BYTES // (1024*1024)}MB zip limit")

    _reap_zip_jobs()
    tmp = tempfile.NamedTemporaryFile(prefix="folderzip-", suffix=".zip", delete=False)
    tmp.close()
    job_id = uuid.uuid4().hex
    job = {
        "status": "running", "error": None, "cancel": False,
        "done": 0, "total": len(files), "bytes_done": 0, "bytes_total": total_bytes,
        "file": tmp.name, "name": (target.name or "projects") + ".zip",
        "created": time.time(),
    }
    ZIP_JOBS[job_id] = job

    def work():
        try:
            with zipfile.ZipFile(tmp.name, "w", zipfile.ZIP_DEFLATED) as zf:
                for fp, size in files:
                    if job["cancel"]:
                        break
                    try:
                        zf.write(fp, str(fp.relative_to(target)))
                        job["bytes_done"] += size
                    except OSError:
                        pass  # file vanished mid-zip; skip it
                    job["done"] += 1
            if job["cancel"]:
                job["status"] = "cancelled"
                os.unlink(tmp.name)
            else:
                job["status"] = "done"
        except Exception as e:  # noqa: BLE001 - report any failure to the poller
            job["status"] = "error"
            job["error"] = str(e)
            try:
                os.unlink(tmp.name)
            except OSError:
                pass

    threading.Thread(target=work, daemon=True).start()
    return {"job": job_id, "total": len(files), "bytes_total": total_bytes}


@app.get("/api/zip/status")
def zip_status(job: str):
    j = ZIP_JOBS.get(job)
    if not j:
        raise HTTPException(404, "Unknown zip job")
    return {k: j[k] for k in ("status", "done", "total", "bytes_done", "bytes_total", "error", "name")}


@app.delete("/api/zip")
def cancel_zip(job: str):
    j = ZIP_JOBS.get(job)
    if not j:
        return {"ok": True}
    if j["status"] == "running":
        j["cancel"] = True
    else:
        try:
            os.unlink(j["file"])
        except OSError:
            pass
        ZIP_JOBS.pop(job, None)
    return {"ok": True}


@app.get("/api/zip/download")
def zip_download(job: str):
    j = ZIP_JOBS.get(job)
    if not j:
        raise HTTPException(404, "Unknown zip job")
    if j["status"] != "done":
        raise HTTPException(409, "Zip is not ready")

    def _cleanup():
        try:
            os.unlink(j["file"])
        except OSError:
            pass
        ZIP_JOBS.pop(job, None)

    return FileResponse(
        j["file"],
        filename=j["name"],
        media_type="application/zip",
        background=BackgroundTask(_cleanup),
    )


@app.get("/api/raw")
def raw_file(path: str):
    """Serve a file inline (Content-Disposition: inline) so PDFs/images render in an iframe."""
    p = safe_path(path)
    if not p.exists() or p.is_dir():
        raise HTTPException(404, "Not found")
    if p.stat().st_size > MAX_TRANSFER_BYTES:
        raise HTTPException(413, f"File exceeds {MAX_TRANSFER_BYTES // (1024*1024)}MB limit")
    mime, _ = mimetypes.guess_type(p.name)
    return FileResponse(
        p,
        media_type=mime or "application/octet-stream",
        content_disposition_type="inline",
    )


def _walk_files(root: Path):
    skip = {".git", "node_modules", ".venv", "venv", "__pycache__", ".mypy_cache", ".pytest_cache", "dist", "build"}
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if d not in skip and not d.startswith(".")]
        for name in filenames:
            yield Path(dirpath) / name


@app.get("/api/search/filename")
def search_filename(q: str = Query(..., min_length=1), limit: int = 50):
    q_lower = q.lower()
    matches = []
    for fp in _walk_files(ROOT):
        if q_lower in fp.name.lower():
            try:
                matches.append(file_info(fp))
            except OSError:
                continue
            if len(matches) >= limit:
                break
    return {"matches": matches}


@app.get("/api/search/content")
def search_content(q: str = Query(..., min_length=2), limit: int = 50):
    q_lower = q.lower()
    matches = []
    for fp in _walk_files(ROOT):
        try:
            st = fp.stat()
        except OSError:
            continue
        if st.st_size > SEARCH_FILE_LIMIT:
            continue
        if not is_text_file(fp):
            continue
        try:
            with fp.open("r", encoding="utf-8", errors="replace") as f:
                for lineno, line in enumerate(f, 1):
                    if q_lower in line.lower():
                        matches.append({
                            "path": to_rel(fp),
                            "name": fp.name,
                            "line": lineno,
                            "snippet": line.strip()[:200],
                        })
                        if len(matches) >= limit:
                            return {"matches": matches}
                        break
        except OSError:
            continue
    return {"matches": matches}


def _load_permlinks() -> list:
    if PERMLINKS_FILE.exists():
        try:
            return json.loads(PERMLINKS_FILE.read_text())
        except (OSError, ValueError):
            return []
    return []


def _save_permlinks(links: list) -> None:
    PERMLINKS_FILE.write_text(json.dumps(links, indent=2))


class PermlinkBody(BaseModel):
    path: str


@app.get("/api/permlinks")
def list_permlinks():
    return {"permlinks": _load_permlinks()}


@app.post("/api/permlinks")
def add_permlink(body: PermlinkBody):
    p = safe_path(body.path)
    if not p.exists() or p.is_dir():
        raise HTTPException(404, "Not found")
    if p.suffix.lower() not in PERMLINK_EXTS:
        raise HTTPException(400, "Only HTML and Markdown files can be perma-linked")
    rel = to_rel(p)
    links = _load_permlinks()
    if not any(l["path"] == rel for l in links):
        links.append({"path": rel, "name": p.name, "created": datetime.now().isoformat()})
        _save_permlinks(links)
    return {"path": rel}


@app.delete("/api/permlinks")
def delete_permlink(path: str):
    rel = to_rel(safe_path(path))
    links = [l for l in _load_permlinks() if l["path"] != rel]
    _save_permlinks(links)
    return {"ok": True}


def _load_bookmarks() -> list:
    if BOOKMARKS_FILE.exists():
        try:
            return json.loads(BOOKMARKS_FILE.read_text())
        except (OSError, ValueError):
            return []
    return []


def _save_bookmarks(marks: list) -> None:
    BOOKMARKS_FILE.write_text(json.dumps(marks, indent=2))


class BookmarkBody(BaseModel):
    path: str


@app.get("/api/bookmarks")
def list_bookmarks():
    return {"bookmarks": _load_bookmarks()}


@app.post("/api/bookmarks")
def add_bookmark(body: BookmarkBody):
    p = safe_path(body.path)
    if not p.exists() or not p.is_dir():
        raise HTTPException(404, "Folder not found")
    rel = to_rel(p)
    name = p.name or "projects"  # root resolves to an empty relative path
    marks = _load_bookmarks()
    if not any(m["path"] == rel for m in marks):
        marks.append({"path": rel, "name": name, "created": datetime.now().isoformat()})
        _save_bookmarks(marks)
    return {"path": rel}


@app.delete("/api/bookmarks")
def delete_bookmark(path: str):
    rel = to_rel(safe_path(path))
    marks = [m for m in _load_bookmarks() if m["path"] != rel]
    _save_bookmarks(marks)
    return {"ok": True}


MD_PAGE_TEMPLATE = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{title}</title>
<script src="https://cdn.jsdelivr.net/npm/marked/marked.min.js"></script>
<style>
  body {{ margin: 0; background: #f8fafc; }}
  .markdown-body {{
    max-width: 880px; margin: 0 auto; padding: 2.5rem 1.25rem;
    color: #1e293b; line-height: 1.65; font-size: 16px;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif; }}
  .markdown-body h1, .markdown-body h2, .markdown-body h3,
  .markdown-body h4, .markdown-body h5, .markdown-body h6 {{
    font-weight: 600; line-height: 1.25; margin: 1.4em 0 .5em; }}
  .markdown-body h1 {{ font-size: 1.9em; border-bottom: 1px solid #e2e8f0; padding-bottom: .3em; }}
  .markdown-body h2 {{ font-size: 1.5em; border-bottom: 1px solid #e2e8f0; padding-bottom: .3em; }}
  .markdown-body h3 {{ font-size: 1.25em; }}
  .markdown-body h4 {{ font-size: 1.05em; }}
  .markdown-body h5, .markdown-body h6 {{ font-size: .95em; color: #475569; }}
  .markdown-body p {{ margin: 0 0 1em; }}
  .markdown-body ul, .markdown-body ol {{ margin: 0 0 1em; padding-left: 1.8em; }}
  .markdown-body ul {{ list-style: disc; }}
  .markdown-body ol {{ list-style: decimal; }}
  .markdown-body li {{ margin: .25em 0; }}
  .markdown-body a {{ color: #2563eb; text-decoration: underline; }}
  .markdown-body strong {{ font-weight: 700; }}
  .markdown-body em {{ font-style: italic; }}
  .markdown-body blockquote {{
    margin: 0 0 1em; padding: .2em 1em; color: #475569;
    border-left: 4px solid #cbd5e1; background: #fff; }}
  .markdown-body code {{
    font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace;
    font-size: .88em; background: #eef2f7; padding: .15em .4em; border-radius: 4px; }}
  .markdown-body pre {{
    background: #f1f5f9; padding: .9em 1em; border-radius: 6px;
    overflow-x: auto; margin: 0 0 1em; }}
  .markdown-body pre code {{ background: none; padding: 0; font-size: .85em; }}
  .markdown-body table {{ border-collapse: collapse; margin: 0 0 1em; }}
  .markdown-body th, .markdown-body td {{ border: 1px solid #cbd5e1; padding: 6px 12px; }}
  .markdown-body th {{ background: #f1f5f9; font-weight: 600; }}
  .markdown-body hr {{ border: 0; border-top: 1px solid #e2e8f0; margin: 1.5em 0; }}
  .markdown-body img {{ max-width: 100%; }}
</style>
</head>
<body>
<div class="markdown-body" id="content"></div>
<script id="md-source" type="text/plain">{source}</script>
<script>
  const src = document.getElementById("md-source").textContent;
  document.getElementById("content").innerHTML = marked.parse(src);
</script>
</body>
</html>"""


def render_markdown_page(name: str, content: str) -> str:
    """Wrap raw markdown in a standalone HTML page that renders it client-side."""
    # Embed the source in a non-executable script block; neutralize any closing
    # tag so the markdown body can't break out of it.
    safe_source = content.replace("</script", "<\\/script")
    return MD_PAGE_TEMPLATE.format(title=html_escape(name), source=safe_source)


@app.get("/view")
def view_html(path: str):
    """Open a file in a new tab: render HTML/Markdown, otherwise serve inline so the
    browser views it (pdf/image) or downloads it (binary office files, etc.)."""
    p = safe_path(path)
    if not p.exists() or p.is_dir():
        raise HTTPException(404, "Not found")
    if p.suffix.lower() in HTML_EXTS:
        return HTMLResponse(p.read_text(encoding="utf-8", errors="replace"))
    if p.suffix.lower() in MD_EXTS:
        return HTMLResponse(render_markdown_page(p.name, p.read_text(encoding="utf-8", errors="replace")))
    if p.stat().st_size > MAX_TRANSFER_BYTES:
        raise HTTPException(413, f"File exceeds {MAX_TRANSFER_BYTES // (1024*1024)}MB limit")
    mime, _ = mimetypes.guess_type(p.name)
    return FileResponse(
        p,
        media_type=mime or "application/octet-stream",
        content_disposition_type="inline",
    )


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


def main() -> None:
    if len(sys.argv) > 1 and sys.argv[1] == "hash-password":
        password = sys.argv[2] if len(sys.argv) > 2 else getpass.getpass("Password: ")
        print(hash_password(password))
        return

    import argparse
    parser = argparse.ArgumentParser(
        prog="filepeek",
        description="Single-file web viewer for your files (md, html, office docs, code).",
        epilog="Other commands:  hash-password [PASSWORD]  — print a FILEPEEK_PASSWORD_HASH value",
    )
    parser.add_argument("--host", default=os.environ.get("FILEPEEK_HOST", "127.0.0.1"))
    parser.add_argument("--port", type=int, default=int(os.environ.get("FILEPEEK_PORT", "8765")))
    args = parser.parse_args()

    if not ROOT.is_dir():
        sys.exit(f"FILEPEEK_ROOT does not exist or is not a directory: {ROOT}")
    if args.host not in ("127.0.0.1", "::1", "localhost") and not AUTH_ENABLED:
        sys.exit(
            f"Refusing to bind {args.host}: no authentication is configured, and this app\n"
            "exposes read/write access to your files. Either bind 127.0.0.1, or set\n"
            "FILEPEEK_PASSWORD_HASH (generate one with: python app.py hash-password)\n"
            "and/or FILEPEEK_TOKEN, then try again."
        )

    import uvicorn
    uvicorn.run(app, host=args.host, port=args.port)


if __name__ == "__main__":
    main()
